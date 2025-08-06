import xml.etree.ElementTree as ET
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os
import datetime
import re

class AllResults:
    re_patter = re.compile("([\\d_]+)[.](\\w+)[.]xml")
    def __init__(self):
        self.all_results = {}

    def load_result(self, platform: str, build_type: str, path: str):
        mx = self.re_patter.match(os.path.basename(path))
        assert(mx), f"Invalid file name format: {path}"
        time = datetime.datetime.strptime(mx.group(1), "%Y_%m_%d_%H_%M_%S")
        hash = mx.group(2)
        if platform not in self.all_results:
            self.all_results[platform] = {}
        self.all_results[platform][build_type] = {}
        self.all_results[platform][build_type]["tests"] = self.parse_ctest_xml(path)
        self.all_results[platform][build_type]["time"] = time
        self.all_results[platform][build_type]["hash"] = hash
        
    def create_spreadsheet(self):
        wb = Workbook()
        for platform, builds in self.all_results.items():
            ws = wb.create_sheet(title=platform)
            tests = sorted(list(set.union(
                *[set(build["tests"].keys()) for build in builds.values()])))
            special = ["time", "hash"]
            
            builds = list(builds.keys())
            header = ["test_name", *builds]
            ws.append(header)
            row_idx = 1 
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            for sp in special:
                ws.append([sp])
                row_idx += 1
                cl = 0
                for build_type in builds:
                    build_results = self.all_results[platform][build_type]
                    res = build_results[sp] if sp in build_results else 'N/A'
                    cl += 1
                    ws[row_idx][cl].value =str(res)
                    ws[row_idx][cl].fill = green_fill if sp in build_results else red_fill
            for test in tests:
                ws.append([test])
                row_idx += 1
                cl = 0
                for build_type in builds:
                    build_results = self.all_results[platform][build_type]["tests"]
                    res = build_results[test]["status"] if test in build_results else 'N/A'
                    color = green_fill if test in build_results \
                      and build_results[test]["failure"] == None else red_fill
                    cl += 1
                    ws[row_idx][cl].value = res
                    ws[row_idx][cl].fill = color


        return wb

    def generate_spreadsheet(self, result_dir, filename: str):
        for plat in os.listdir(result_dir):
            pd = os.path.join(result_dir, plat)
            for build_type in os.listdir(pd):
                bd = os.path.join(pd, build_type)
                file_name = sorted([x for x in os.listdir(bd) if x.endswith('.xml')])[-1] 
                self.load_result(plat, build_type, os.path.join(bd, file_name))

        wb = self.create_spreadsheet()
        wb.save(filename)

    @staticmethod
    def parse_ctest_xml(xml_path):
        tree = ET.parse(xml_path)
        root = tree.getroot()
        results = {}
        for testcase in root.findall('testcase'):
            name = testcase.get('name')
            classname = testcase.get('classname')
            time = testcase.get('time')
            status = testcase.get('status')
            failure = testcase.find('failure') is not None
            results[name] = {
                'classname': classname,
                'time': time,
                'status': status,
                'failure': failure
            }
        return results
# all_results["platform"]["build_type"]


def main():
    all_results = AllResults()
    all_results.generate_spreadsheet('/home/zmartonka/test_results', './ctest_results.xlsx')
    
if __name__ == '__main__':
    main()
